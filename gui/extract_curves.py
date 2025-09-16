import openpyxl
import pandas as pd
import numpy as np
from os import path
from glob import glob
import matplotlib
import matplotlib.pyplot as plt 
from functions.functions import calculate_slopes

global resulting_dfs
output = {}
resulting_dfs = []


def dodaj_prady(cycle, column_to_average = 'Im_GEO'):
    sciezka = path.abspath(input('Podaj ścieżke: '))
    files = glob(sciezka +'/*.xlsx')
    result = []
    cycles = list(cycle)
    for cycle in cycles:
        dfs = []
        for file in files:
            wb = openpyxl.load_workbook(path.abspath(file))
            lsv1 = wb[f'CYCLE {str(cycle)}_LSV']
            print(lsv1)

            tmp = []
            for column in lsv1.iter_cols():
                
                column_name = column[0].value
                if column_name == column_to_average:
                    j_geo = []
                    for cell in column[1:]:
                        j_geo.append(cell.value)
                    j_geo = np.array(j_geo, dtype=np.float64)
                    tmp.append(j_geo)
            df = pd.DataFrame(tmp).T
            dfs.append(df)
        dfs = pd.concat(dfs,axis=1)
        dfs['mean'] = dfs.mean(axis=1)
        dfs['stdev'] = dfs.std(axis=1)

        result.append(dfs[['mean','stdev']])

    return result





def tafel(cycle, sciezka, klucz=None, mode = 'GEO'):

    if mode == 'GEO':
        current_column = 'log10 Im_GEO'
    elif mode == 'ECSA':
        current_column = 'log10 Im_ECSA'
    
    if klucz is not None:
        files = glob(f'{sciezka}/*{klucz}*.xlsx')
    dfs = []
    for file in files:
        i = 0
        wb = openpyxl.load_workbook(path.abspath(file))
        lsv1 = wb[f'CYCLE {str(cycle)}_TAFEL']
        print(lsv1)
        tmp = []
        vf_ir, j = [], [ ]
        result_df = []
        for column_number, column in enumerate(lsv1.iter_cols()):
            column_name = column[0].value

            if column_name == current_column:
                j = []
                for cell in column[1:]:
                    j.append(cell.value)
                j = np.array(j, dtype=np.float64)
                continue
            elif column_name == 'Vf_iR':
                vf_ir = []
                for cell in column[1:]:
                    vf_ir.append(cell.value)
                vf_ir = np.array(vf_ir, dtype=np.float64)
   
                if len(vf_ir) > 0 and len(j) > 0:
                    df_tmp = pd.DataFrame({'Vf_iR':vf_ir, 'log10 Im_GEO':j})
                    print(df_tmp)
                    result = calculate_slopes(df_tmp, -25, -5, -2.5, name = path.basename(file), curve_number = i)
        
                    vf_ir = []
                    j = []
                    i += 1

                #fig, ax = plt.subplots()
                #line, = ax.plot(vf_ir, j_geo)
                #selected_point, = ax.plot([], [], 'ro')
                #plt.title(f"test")
                #plt.xlabel("Vf_iR")
                #plt.ylabel("j_GEO")
                #ax.legend()
                #plt.show()
    return 'ECSA_' +path.basename(file)



#def process_tafels():
klucz = '2RU'
normalization = 'ECSA'
sciezka = path.abspath(input('Podaj ścieżke: '))
tafel(1,sciezka, mode = normalization)
name = 'TAFEL'


resulting_dfs = pd.concat(resulting_dfs, axis=1)


try: 
    with pd.ExcelWriter(sciezka+f'/{name}.xlsx', mode = 'w') as writer:
        resulting_dfs.to_excel(writer, engine='openpyxl',sheet_name='cycle1')
except:
    with pd.ExcelWriter(sciezka+f'/{name}.xlsx', mode = 'a') as writer:
        resulting_dfs.to_excel(writer, engine='openpyxl',sheet_name='cycle1')


resulting_dfs = []
tafel(2,sciezka, mode = normalization)
resulting_dfs = pd.concat(resulting_dfs, axis=1)

try: 
    with pd.ExcelWriter(sciezka+f'/{name}.xlsx', mode = 'a') as writer:
        resulting_dfs.to_excel(writer, engine='openpyxl',sheet_name='cycle2')
except:
    with pd.ExcelWriter(sciezka+'/tafel2.xlsx', mode = 'w') as writer:
        resulting_dfs.to_excel(writer, engine='openpyxl',sheet_name='cycle2')
