import openpyxl
import pandas as pd
import numpy as np
from os import path
from glob import glob
import matplotlib
import matplotlib.pyplot as plt 

global resulting_dfs
output = {}
resulting_dfs = []


def dodaj_prady(cycle, column_to_average = 'Im_GEO'):
    sciezka = path.abspath(input('Podaj ścieżke: '))
    files = glob(sciezka +'/*.xlsx')
    result = []
    cycles = list(cycle)
    for cycle in cycles:
        dfs = []
        for file in files:
            wb = openpyxl.load_workbook(path.abspath(file))
            lsv1 = wb[f'CYCLE {str(cycle)}_LSV']
            print(lsv1)

            tmp = []
            for column in lsv1.iter_cols():
                
                column_name = column[0].value
                if column_name == column_to_average:
                    j_geo = []
                    for cell in column[1:]:
                        j_geo.append(cell.value)
                    j_geo = np.array(j_geo, dtype=np.float64)
                    tmp.append(j_geo)
            df = pd.DataFrame(tmp).T
            dfs.append(df)
        dfs = pd.concat(dfs,axis=1)
        dfs['mean'] = dfs.mean(axis=1)
        dfs['stdev'] = dfs.std(axis=1)

        result.append(dfs[['mean','stdev']])

    return result


def calculate_slope(data, starting_point, step, overlap, name = 'Sample', curve_number = None, i = None):

    if i == None:
        output[curve_number] = []

    i_start = (np.abs(data['Vf_iR'] - starting_point)).argmin()
    print("Starting index: ", i_start)
    search = data['Vf_iR'][i_start]
    print("Potential of starting point: ", search)
    new_search = search + step
    print("Potential of new point: ", new_search)
    idx = (np.abs(data['Vf_iR'] - new_search)).argmin()
    print("Finishing index: ", idx)
    if (idx < i_start) or (new_search < min(data['Vf_iR'])):

        fig, ax = plt.subplots(figsize=(15,10))
        plt.xlabel('Average log10 j [mA/cm2]')
        plt.ylabel('Tafel slope [mV/dec]')
        plt.title(name)
        
        print('Calculated finish index is lower than starting index. Aborting. This is due to iR-drop and bubbles detachment.')
        df = pd.DataFrame(output[curve_number], columns = ['E_begining', 'E_final', 'Average current [mA/cm2]', 'Tafel slope [mV/dec]'])
        resulting_dfs.append(df)

        x_data = np.array([seg[2] for seg in output[curve_number]])
        y_data = np.array([seg[3] for seg in output[curve_number]])
        ax.scatter(x_data,y_data)
        ax.set_ylim(0,150)
        clicked_points = []

        
        def on_click(event):
            if event.inaxes != ax:
                return
            
            clicked_points.append((event.xdata, event.ydata))

            # Draw a red dot
            ax.plot(event.xdata, event.ydata, 'ro')
            fig.canvas.draw()

            if len(clicked_points) == 2:
                # Get x-values of clicks
                x1, _ = clicked_points[0]
                x2, _ = clicked_points[1]

                # Find closest indices
                idx1 = (np.abs(x_data - x1)).argmin()
                idx2 = (np.abs(x_data - x2)).argmin()

                i_min, i_max = sorted([idx1, idx2])  # Ensure correct order

                selected_x = x_data[i_min:i_max+1]
                selected_y = y_data[i_min:i_max+1]

                mean_val = np.mean(selected_y)
                print(f"\nSelected range: x = [{x_data[i_min]:.3f}, {x_data[i_max]:.3f}]")
                print(f"Mean y-value in range: {mean_val:.5f}")

                # Optionally shade selected region
                ax.axvspan(x_data[i_min], x_data[i_max], color='orange', alpha=0.3)
                fig.canvas.draw()

                # Disconnect listener so it doesn’t keep listening
                fig.canvas.mpl_disconnect(cid)
                plt.title(f"Mean y = {mean_val:.5f} between selected points")
                fig.canvas.draw()

        # Connect the click event
        cid = fig.canvas.mpl_connect('button_press_event', on_click)
        plt.show()
        

        return df

    new_search_overlap = new_search - overlap
    print('Next potential, including overlap: ', new_search_overlap)
    print(new_search)
    try:
        x = data['log10 Im_GEO'][i_start:idx]
    except:
        x = data['log10 Im_ECSA'][i_start:idx]
    y = data['Vf_iR'][i_start:idx]
    slope, intercept = np.polyfit(x, -y, 1)
    average_current = np.mean(x)
    #ax.scatter(average_current, slope)

    output[curve_number].append((search, new_search, average_current, slope))

    if min(data['Vf_iR']) < new_search:
        
        calculate_slope(data, new_search_overlap, step, overlap = overlap, name = name, curve_number = curve_number, i = 1)
    else:
        
        #plt.show()
        return


def tafel(cycle, sciezka, klucz=None, mode = 'GEO'):

    if mode == 'GEO':
        current_column = 'log10 Im_GEO'
    elif mode == 'ECSA':
        current_column = 'log10 Im_ECSA'
    
    files = glob(f'{sciezka}/*{klucz}*.xlsx')
    dfs = []
    for file in files:
        i = 0
        wb = openpyxl.load_workbook(path.abspath(file))
        lsv1 = wb[f'CYCLE {str(cycle)}_TAFEL']
        print(lsv1)
        tmp = []
        vf_ir, j = [], [ ]
        result_df = []
        for column_number, column in enumerate(lsv1.iter_cols()):
            column_name = column[0].value

            if column_name == current_column:
                j = []
                for cell in column[1:]:
                    j.append(cell.value)
                j = np.array(j, dtype=np.float64)
                continue
            elif column_name == 'Vf_iR':
                vf_ir = []
                for cell in column[1:]:
                    vf_ir.append(cell.value)
                vf_ir = np.array(vf_ir, dtype=np.float64)
   
                if len(vf_ir) > 0 and len(j) > 0:
                    df_tmp = pd.DataFrame({'Vf_iR':vf_ir, 'log10 Im_GEO':j})
                    print(df_tmp)
                    result = calculate_slope(df_tmp, -25, -5, -2.5, name = path.basename(file), curve_number = i)
        
                    vf_ir = []
                    j = []
                    i += 1

                #fig, ax = plt.subplots()
                #line, = ax.plot(vf_ir, j_geo)
                #selected_point, = ax.plot([], [], 'ro')
                #plt.title(f"test")
                #plt.xlabel("Vf_iR")
                #plt.ylabel("j_GEO")
                #ax.legend()
                #plt.show()
    return 'ECSA_' +path.basename(file)



#def process_tafels():
klucz = '2RU'
normalization = 'ECSA'
sciezka = path.abspath(input('Podaj ścieżke: '))
tafel(1,sciezka, klucz, mode = normalization)
name = 'TAFEL'


resulting_dfs = pd.concat(resulting_dfs, axis=1)


try: 
    with pd.ExcelWriter(sciezka+f'/{name}.xlsx', mode = 'w') as writer:
        resulting_dfs.to_excel(writer, engine='openpyxl',sheet_name='cycle1')
except:
    with pd.ExcelWriter(sciezka+f'/{name}.xlsx', mode = 'a') as writer:
        resulting_dfs.to_excel(writer, engine='openpyxl',sheet_name='cycle1')


resulting_dfs = []
tafel(2,sciezka, klucz, mode = normalization)
resulting_dfs = pd.concat(resulting_dfs, axis=1)

try: 
    with pd.ExcelWriter(sciezka+f'/{name}.xlsx', mode = 'a') as writer:
        resulting_dfs.to_excel(writer, engine='openpyxl',sheet_name='cycle2')
except:
    with pd.ExcelWriter(sciezka+'/tafel2.xlsx', mode = 'w') as writer:
        resulting_dfs.to_excel(writer, engine='openpyxl',sheet_name='cycle2')
